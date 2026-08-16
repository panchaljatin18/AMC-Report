import openpyxl
from pathlib import Path
from typing import List, Dict, Any

def _safe_float(val, default=0.0) -> float:
    if val is None:
        return default
    if isinstance(val, (int, float)):
        return float(val)
    val_str = str(val).replace(",", "").strip()
    if not val_str or val_str.lower() in ["none", "nan", "n/a", "-", "wip"]:
        return default
    try:
        return float(val_str)
    except (ValueError, TypeError):
        return default

def _safe_int_or_float(val: float):
    return int(val) if val.is_integer() else val

def parse_excel_fast(file_path: str, target_sheets: list):
    wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
    target_sheet_name = None
    for s in wb.sheetnames:
        if s.strip().lower() in [ts.lower() for ts in target_sheets]:
            target_sheet_name = s
            break
    if not target_sheet_name:
        target_sheet_name = wb.sheetnames[0]

    sheet = wb[target_sheet_name]
    all_rows = []
    for row in sheet.iter_rows(values_only=True):
        all_rows.append(list(row))
    wb.close()
    return all_rows, target_sheet_name

def parse_road_data(all_rows):
    header_idx = -1
    for r_idx in range(min(15, len(all_rows))):
        row_vals = [str(x).strip().lower() for x in all_rows[r_idx] if x is not None]
        if "zone" in row_vals and ("problem" in row_vals or "open" in row_vals or "grand total" in row_vals):
            header_idx = r_idx
            break

    if header_idx == -1:
        return []

    headers = [str(x).strip() if x is not None else f"col_{i}" for i, x in enumerate(all_rows[header_idx])]
    
    col_zone = next((i for i, c in enumerate(headers) if "zone" in c.lower()), 0)
    col_prob = next((i for i, c in enumerate(headers) if "problem" in c.lower()), 1 if len(headers) > 1 else col_zone)
    col_closed = next((i for i, c in enumerate(headers) if "closed" in c.lower()), None)
    col_open = next((i for i, c in enumerate(headers) if "open" in c.lower() and "grand" not in c.lower()), None)
    col_gt = next((i for i, c in enumerate(headers) if "grand total" in c.lower() or "total" in c.lower()), None)

    current_zone = ""
    rows = []

    for r_idx in range(header_idx + 1, len(all_rows)):
        row = all_rows[r_idx]
        raw_z = str(row[col_zone]).strip() if col_zone < len(row) and row[col_zone] is not None else ""
        raw_p = str(row[col_prob]).strip() if col_prob < len(row) and row[col_prob] is not None else ""

        if raw_z.lower().startswith("zone wise") or raw_z.lower().startswith("reporting") or "total" in raw_z.lower() or raw_z.lower() in ["zone", "row labels"]:
            if "total" in raw_z.lower():
                current_zone = ""
            continue

        if raw_z != "" and not raw_z.lower().endswith("total"):
            current_zone = raw_z

        if not raw_p or raw_p.lower() in ["problem", "row labels"] or "total" in raw_p.lower():
            continue

        if not current_zone:
            continue

        closed = _safe_float(row[col_closed]) if col_closed is not None and col_closed < len(row) else 0.0
        open_cnt = _safe_float(row[col_open]) if col_open is not None and col_open < len(row) else 0.0
        gt = _safe_float(row[col_gt]) if col_gt is not None and col_gt < len(row) else closed + open_cnt

        rows.append({
            "row_index": r_idx + 1,
            "zone": current_zone,
            "problem": raw_p,
            "closed": _safe_int_or_float(closed),
            "open": _safe_int_or_float(open_cnt),
            "grand_total": _safe_int_or_float(gt)
        })

    return rows

def parse_water_data(all_rows):
    header_idx = -1
    for r_idx in range(min(15, len(all_rows))):
        row_vals = [str(x).strip().lower() for x in all_rows[r_idx] if x is not None]
        if "zone" in row_vals and ("problem" in row_vals or "open" in row_vals):
            header_idx = r_idx
            break

    if header_idx != -1:
        headers = [str(x).strip() if x is not None else f"col_{i}" for i, x in enumerate(all_rows[header_idx])]
        col_zone = next((i for i, c in enumerate(headers) if "zone" in c.lower()), 0)
        col_prob = next((i for i, c in enumerate(headers) if "problem" in c.lower()), 1 if len(headers) > 1 else col_zone)
        col_open = next((i for i, c in enumerate(headers) if "open" in c.lower() and "grand" not in c.lower()), None)
        if col_open is None:
            col_open = next((i for i, c in enumerate(headers) if "open" in c.lower()), None)

        zone_categories = {}
        all_categories = set()
        current_zone = ""

        for r_idx in range(header_idx + 1, len(all_rows)):
            row = all_rows[r_idx]
            raw_z = str(row[col_zone]).strip() if col_zone < len(row) and row[col_zone] is not None else ""
            raw_p = str(row[col_prob]).strip() if col_prob < len(row) and row[col_prob] is not None else ""

            if raw_z.lower().startswith("zone wise") or "total" in raw_z.lower() or raw_z.lower() in ["zone", "row labels"]:
                if "total" in raw_z.lower():
                    current_zone = ""
                continue

            if raw_z != "" and not raw_z.lower().endswith("total"):
                current_zone = raw_z

            if not raw_p or raw_p.lower() in ["problem", "row labels"] or "total" in raw_p.lower():
                continue

            if not current_zone:
                continue

            cat_name = raw_p
            for prefix in ["Water-", "Water -", "Water "]:
                if cat_name.startswith(prefix):
                    cat_name = cat_name[len(prefix):].strip()

            open_val = _safe_float(row[col_open]) if col_open is not None and col_open < len(row) else 0.0

            if current_zone not in zone_categories:
                zone_categories[current_zone] = {}
            zone_categories[current_zone][cat_name] = zone_categories[current_zone].get(cat_name, 0.0) + open_val
            all_categories.add(cat_name)

        matrix_rows = []
        for z, cats in zone_categories.items():
            r_data = {"zone": z}
            tot_open = 0.0
            for cat in sorted(list(all_categories)):
                val = cats.get(cat, 0.0)
                int_val = _safe_int_or_float(val)
                r_data[cat] = int_val
                tot_open += val
            r_data["Total Open"] = _safe_int_or_float(tot_open)
            matrix_rows.append(r_data)
        return matrix_rows

    # Fallback to pivot / matrix format
    pivot_header_idx = -1
    for r_idx in range(min(15, len(all_rows))):
        row_vals = [str(x).strip().lower() for x in all_rows[r_idx] if x is not None]
        if "row labels" in row_vals or "closed" in row_vals or "open" in row_vals:
            pivot_header_idx = r_idx
            break

    if pivot_header_idx == -1:
        return []

    headers = [str(x).strip() if x is not None else f"col_{i}" for i, x in enumerate(all_rows[pivot_header_idx])]
    col_label = 0
    for i, c in enumerate(headers):
        if c.lower() in ["row labels", "zone", "problem"]:
            col_label = i
            break
    col_open = next((i for i, c in enumerate(headers) if "open" in c.lower() and "re" not in c.lower() and "grand" not in c.lower()), 1)

    zone_categories = {}
    all_categories = set()
    current_zone = ""

    for r_idx in range(pivot_header_idx + 1, len(all_rows)):
        row = all_rows[r_idx]
        lbl = str(row[col_label]).strip() if col_label < len(row) and row[col_label] is not None else ""
        if not lbl or lbl.lower().startswith("zone wise") or "total" in lbl.lower() or lbl.lower() in ["row labels", "zone"]:
            continue

        if lbl.isupper() and len(lbl) < 20 and not lbl.startswith("Water"):
            current_zone = lbl
            continue

        if not current_zone:
            continue

        cat_name = lbl
        for prefix in ["Water-", "Water -", "Water "]:
            if cat_name.startswith(prefix):
                cat_name = cat_name[len(prefix):].strip()

        open_val = _safe_float(row[col_open]) if col_open < len(row) else 0.0
        if current_zone not in zone_categories:
            zone_categories[current_zone] = {}
        zone_categories[current_zone][cat_name] = zone_categories[current_zone].get(cat_name, 0.0) + open_val
        all_categories.add(cat_name)

    matrix_rows = []
    for z, cats in zone_categories.items():
        r_data = {"zone": z}
        tot_open = 0.0
        for cat in sorted(list(all_categories)):
            val = cats.get(cat, 0.0)
            int_val = _safe_int_or_float(val)
            r_data[cat] = int_val
            tot_open += val
        r_data["Total Open"] = _safe_int_or_float(tot_open)
        matrix_rows.append(r_data)
    return matrix_rows


upload_dir = Path("c:/Users/Jmpan/OneDrive/Desktop/AMC CRM/backend/uploads/bedef3ac")
f_road = str(upload_dir / "11-08-2026 2 PM CCRS REPORT-Road.xlsx")
f_drainage = str(upload_dir / "11-08-2026 2 PM CCRS REPORT-Drainage.xlsx")
f_water = str(upload_dir / "11-08-2026 2 PM CCRS REPORT-Water.xlsx")

print("--- TESTING ROAD ---")
road_rows_raw, r_sheet = parse_excel_fast(f_road, ["road 1", "road"])
road_rows = parse_road_data(road_rows_raw)
print(f"Sheet '{r_sheet}' -> Parsed {len(road_rows)} rows:")
for r in road_rows[:5]:
    print(" ", r)

print("\n--- TESTING DRAINAGE ---")
drainage_rows_raw, d_sheet = parse_excel_fast(f_drainage, ["DRAINAGE", "drainage 1"])
drainage_rows = parse_road_data(drainage_rows_raw)
print(f"Sheet '{d_sheet}' -> Parsed {len(drainage_rows)} rows:")
for r in drainage_rows[:5]:
    print(" ", r)

print("\n--- TESTING WATER ---")
water_rows_raw, w_sheet = parse_excel_fast(f_water, ["WATER", "water1"])
water_rows = parse_water_data(water_rows_raw)
print(f"Sheet '{w_sheet}' -> Parsed {len(water_rows)} matrix rows:")
for r in water_rows[:5]:
    print(" ", r)
