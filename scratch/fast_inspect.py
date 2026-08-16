import openpyxl

wb = openpyxl.load_workbook("c:/Users/Jmpan/OneDrive/Desktop/AMC CRM/backend/uploads/bedef3ac/11-08-2026 2 PM CCRS REPORT-Water.xlsx", read_only=True, data_only=True)
print("Sheet names:", wb.sheetnames)

sheet_name = "water1" if "water1" in wb.sheetnames else wb.sheetnames[0]
sheet = wb[sheet_name]

print(f"\n--- Top 20 rows of '{sheet_name}' ---")
for i, row in enumerate(sheet.iter_rows(values_only=True)):
    if i >= 25:
        break
    row_filtered = [v for v in row if v is not None]
    if row_filtered:
        print(f"Row {i:2d}:", row_filtered[:10])
