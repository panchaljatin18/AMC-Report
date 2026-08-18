import openpyxl
from typing import List, Dict, Any

def _safe_float(val, default=0.0) -> float:
    if val is None:
        return default
    if isinstance(val, (int, float)):
        return float(val)
    val_str = str(val).replace(",", "").strip()
    if not val_str or val_str.lower() in ["none", "nan", "n/a", "-", "wip"]:
        return default
    try:
        return float(val_str)
    except (ValueError, TypeError):
        return default

def _safe_int_or_float(val: float):
    return int(val) if val.is_integer() else val

def _clean_category_name(name: str, domain: str = "") -> str:
    name = str(name).strip()
    if domain.lower() == "drainage":
        for prefix in ["Drainage-", "Drainage - ", "Drainage ", "DRAINAGE-", "DRAINAGE - "]:
            if name.startswith(prefix):
                name = name[len(prefix):].strip()
        if "public toilets" in name.lower() or "urinals" in name.lower():
            return "Public Toilets/Urinals"
        if "choking" in name.lower() and "line" in name.lower():
            return "Choking Of Line"
        if "manhole" in name.lower():
            return "Manhole Cover Missing"
        if name.lower() in ["other", "others"]:
            return "Other"
    elif domain.lower() == "road":
        for prefix in ["Road-", "Road - ", "Road ", "ROAD-", "ROAD - "]:
            if name.startswith(prefix):
                name = name[len(prefix):].strip()
    elif domain.lower() == "water":
        for prefix in ["Water-", "Water - ", "Water ", "WATER-", "WATER - "]:
            if name.startswith(prefix):
                name = name[len(prefix):].strip()
    return name

def _read_sheet_rows_fast(file_path: str, target_sheet_candidates: List[str]):
    wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
    target_sheet_name = None

    # Priority 1: Exact match in candidate priority order
    for cand in target_sheet_candidates:
        for s in wb.sheetnames:
            if s.strip().lower() == cand.strip().lower():
                target_sheet_name = s
                break
        if target_sheet_name:
            break

    # Priority 2: Substring match in candidate priority order
    if not target_sheet_name:
        for cand in target_sheet_candidates:
            for s in wb.sheetnames:
                if cand.strip().lower() in s.strip().lower():
                    target_sheet_name = s
                    break
            if target_sheet_name:
                break

    # Priority 3: First sheet fallback
    if not target_sheet_name:
        target_sheet_name = wb.sheetnames[0]

    sheet = wb[target_sheet_name]
    all_rows = []
    for row in sheet.iter_rows(values_only=True):
        if row and any(x is not None for x in row):
            all_rows.append(list(row))
    wb.close()
    return all_rows


def parse_road_excel(file_path: str) -> List[Dict[str, Any]]:
    """
    Parses Road.xlsx sheet 'road' or 'ROAD'
    Supports standard tables, pivot tables, merged zone cells, and subtotal rows.
    """
    try:
        all_rows = _read_sheet_rows_fast(file_path, ["road", "ROAD", "Road", "road 1", "3 CATEGORY", "Sheet1"])
        header_idx = -1
        for r_idx in range(min(25, len(all_rows))):
            row_vals = [str(x).strip().lower() for x in all_rows[r_idx] if x is not None]
            has_zone = any(k in row_vals for k in ["zone", "zone name", "row labels", "zones"])
            has_metric = any(k in row_vals for k in ["open", "closed", "grand total", "total", "problem", "repair"])
            if has_zone and has_metric:
                header_idx = r_idx
                break

        if header_idx == -1:
            return []

        raw_header = all_rows[header_idx]
        first_non_empty = next((i for i, x in enumerate(raw_header) if x is not None), 0)
        headers = [str(x).strip() if x is not None else f"col_{i}" for i, x in enumerate(raw_header[first_non_empty:], start=first_non_empty)]
        
        col_zone = next((i for i, c in enumerate(headers) if any(k in c.lower() for k in ["zone", "row labels"])), 0)
        col_prob = next((i for i, c in enumerate(headers) if "problem" in c.lower() or "category" in c.lower() or "nature" in c.lower()), None)
        col_closed = next((i for i, c in enumerate(headers) if "closed" in c.lower()), None)
        col_open = next((i for i, c in enumerate(headers) if "open" in c.lower() and "re" not in c.lower() and "grand" not in c.lower()), None)
        col_gt = next((i for i, c in enumerate(headers) if "grand total" in c.lower() or "total" in c.lower()), None)

        current_zone = ""
        rows = []

        for r_idx in range(header_idx + 1, len(all_rows)):
            raw_row = all_rows[r_idx]
            if not raw_row:
                continue
            row = raw_row[first_non_empty:]
            if not row:
                continue

            raw_z = str(row[col_zone]).strip() if col_zone < len(row) and row[col_zone] is not None else ""
            raw_p = str(row[col_prob]).strip() if col_prob is not None and col_prob < len(row) and row[col_prob] is not None else ""

            if raw_z.lower().startswith("zone wise") or raw_z.lower().startswith("reporting") or "total" in raw_z.lower() or raw_z.lower() in ["zone", "row labels", "zone name"]:
                if "total" in raw_z.lower():
                    current_zone = ""
                continue

            # Case 1: Separate zone and problem columns
            if col_prob is not None and col_prob != col_zone:
                if raw_z != "" and not raw_z.lower().endswith("total"):
                    current_zone = raw_z
                if not raw_p or raw_p.lower() in ["problem", "row labels"] or "total" in raw_p.lower():
                    continue
                if not current_zone:
                    continue

                closed = _safe_float(row[col_closed]) if col_closed is not None and col_closed < len(row) else 0.0
                open_cnt = _safe_float(row[col_open]) if col_open is not None and col_open < len(row) else 0.0
                gt = _safe_float(row[col_gt]) if col_gt is not None and col_gt < len(row) else closed + open_cnt

                rows.append({
                    "row_index": r_idx + 1,
                    "zone": current_zone,
                    "problem": _clean_category_name(raw_p, "road"),
                    "closed": _safe_int_or_float(closed),
                    "open": _safe_int_or_float(open_cnt),
                    "grand_total": _safe_int_or_float(gt)
                })
            # Case 2: Pivot format
            else:
                if not raw_z:
                    continue
                if raw_z.isupper() and len(raw_z) < 25 and not raw_z.startswith("ROAD"):
                    current_zone = raw_z
                    continue
                if not current_zone:
                    continue

                closed = _safe_float(row[col_closed]) if col_closed is not None and col_closed < len(row) else 0.0
                open_cnt = _safe_float(row[col_open]) if col_open is not None and col_open < len(row) else 0.0
                gt = _safe_float(row[col_gt]) if col_gt is not None and col_gt < len(row) else closed + open_cnt

                rows.append({
                    "row_index": r_idx + 1,
                    "zone": current_zone,
                    "problem": _clean_category_name(raw_z, "road"),
                    "closed": _safe_int_or_float(closed),
                    "open": _safe_int_or_float(open_cnt),
                    "grand_total": _safe_int_or_float(gt)
                })

        return rows
    except Exception as e:
        raise ValueError(f"Error parsing Road Excel file: {str(e)}")


def parse_drainage_excel(file_path: str) -> List[Dict[str, Any]]:
    """
    Parses Drainage.xlsx sheet 'DRAINAGE', 'Drainage', 'drainage 1', 'pivot'
    Supports standard tables, pivot tables, merged zone cells, and subtotal rows.
    """
    try:
        all_rows = _read_sheet_rows_fast(file_path, ["DRAINAGE", "Drainage", "drainage 1", "pivot", "Sheet1"])
        header_idx = -1
        for r_idx in range(min(25, len(all_rows))):
            row_vals = [str(x).strip().lower() for x in all_rows[r_idx] if x is not None]
            has_zone = any(k in row_vals for k in ["zone", "zone name", "row labels", "zones"])
            has_metric = any(k in row_vals for k in ["open", "closed", "grand total", "total", "problem", "choking"])
            if has_zone and has_metric:
                header_idx = r_idx
                break

        if header_idx == -1:
            return []

        raw_header = all_rows[header_idx]
        first_non_empty = next((i for i, x in enumerate(raw_header) if x is not None), 0)
        headers = [str(x).strip() if x is not None else f"col_{i}" for i, x in enumerate(raw_header[first_non_empty:], start=first_non_empty)]
        
        col_zone = next((i for i, c in enumerate(headers) if any(k in c.lower() for k in ["zone", "row labels"])), 0)
        col_prob = next((i for i, c in enumerate(headers) if "problem" in c.lower() or "category" in c.lower() or "nature" in c.lower()), None)
        col_closed = next((i for i, c in enumerate(headers) if "closed" in c.lower()), None)
        col_open = next((i for i, c in enumerate(headers) if "open" in c.lower() and "re" not in c.lower() and "grand" not in c.lower()), None)
        col_gt = next((i for i, c in enumerate(headers) if "grand total" in c.lower() or "total" in c.lower()), None)

        current_zone = ""
        rows = []

        for r_idx in range(header_idx + 1, len(all_rows)):
            raw_row = all_rows[r_idx]
            if not raw_row:
                continue
            row = raw_row[first_non_empty:]
            if not row:
                continue

            raw_z = str(row[col_zone]).strip() if col_zone < len(row) and row[col_zone] is not None else ""
            raw_p = str(row[col_prob]).strip() if col_prob is not None and col_prob < len(row) and row[col_prob] is not None else ""

            if raw_z.lower().startswith("zone wise") or raw_z.lower().startswith("reporting") or "total" in raw_z.lower() or raw_z.lower() in ["zone", "row labels", "zone name"]:
                if "total" in raw_z.lower():
                    current_zone = ""
                continue

            # Case 1: Separate zone and problem columns
            if col_prob is not None and col_prob != col_zone:
                if raw_z != "" and not raw_z.lower().endswith("total"):
                    current_zone = raw_z
                if not raw_p or raw_p.lower() in ["problem", "row labels"] or "total" in raw_p.lower():
                    continue
                if not current_zone:
                    continue

                closed = _safe_float(row[col_closed]) if col_closed is not None and col_closed < len(row) else 0.0
                open_cnt = _safe_float(row[col_open]) if col_open is not None and col_open < len(row) else 0.0
                gt = _safe_float(row[col_gt]) if col_gt is not None and col_gt < len(row) else closed + open_cnt

                rows.append({
                    "row_index": r_idx + 1,
                    "zone": current_zone,
                    "problem": _clean_category_name(raw_p, "drainage"),
                    "closed": _safe_int_or_float(closed),
                    "open": _safe_int_or_float(open_cnt),
                    "grand_total": _safe_int_or_float(gt)
                })
            # Case 2: Pivot format (Zone and Problem in same column)
            else:
                if not raw_z:
                    continue
                if raw_z.isupper() and len(raw_z) < 25 and not raw_z.startswith("DRAINAGE"):
                    current_zone = raw_z
                    continue
                if not current_zone:
                    continue

                closed = _safe_float(row[col_closed]) if col_closed is not None and col_closed < len(row) else 0.0
                open_cnt = _safe_float(row[col_open]) if col_open is not None and col_open < len(row) else 0.0
                gt = _safe_float(row[col_gt]) if col_gt is not None and col_gt < len(row) else closed + open_cnt

                rows.append({
                    "row_index": r_idx + 1,
                    "zone": current_zone,
                    "problem": _clean_category_name(raw_z, "drainage"),
                    "closed": _safe_int_or_float(closed),
                    "open": _safe_int_or_float(open_cnt),
                    "grand_total": _safe_int_or_float(gt)
                })

        return rows
    except Exception as e:
        raise ValueError(f"Error parsing Drainage Excel file: {str(e)}")


def parse_water_excel(file_path: str) -> List[Dict[str, Any]]:
    """
    Parses Water.xlsx sheet 'WATER', 'water', 'water1'
    Supports:
    1. Direct Matrix format: Zone | Category1 | Category2 | ... | Total Open
    2. Row-wise format: Zone | Problem | Closed | Open
    3. Pivot format: Row Labels | Closed | Open
    """
    try:
        all_rows = _read_sheet_rows_fast(file_path, ["WATER", "water", "water1", "Sheet1"])
        header_idx = -1
        for r_idx in range(min(25, len(all_rows))):
            row_vals = [str(x).strip().lower() for x in all_rows[r_idx] if x is not None]
            has_zone_header = "zone" in row_vals or "zone name" in row_vals or "zones" in row_vals
            has_matrix_header = any(k in row_vals for k in ["no supply", "leakage", "pollution", "low pressure", "tanker", "other"])
            has_pivot_header = "row labels" in row_vals
            if has_zone_header or has_matrix_header or has_pivot_header:
                header_idx = r_idx
                break

        if header_idx == -1:
            return []

        raw_header = all_rows[header_idx]
        first_non_empty = next((i for i, x in enumerate(raw_header) if x is not None), 0)
        headers = [str(x).strip() if x is not None else f"col_{i}" for i, x in enumerate(raw_header[first_non_empty:], start=first_non_empty)]
        headers_lower = [h.lower() for h in headers]

        has_problem_col = any("problem" in h or "category" in h or "nature" in h for h in headers_lower)
        has_matrix_cols = any(h in ["no supply", "leakage", "pollution", "low pressure", "tanker", "other"] for h in headers_lower)

        # 1. Direct Matrix format
        if has_matrix_cols and not has_problem_col:
            col_zone = next((i for i, h in enumerate(headers) if "zone" in h.lower()), 0)
            matrix_rows = []
            for r_idx in range(header_idx + 1, len(all_rows)):
                raw_row = all_rows[r_idx]
                if not raw_row:
                    continue
                row = raw_row[first_non_empty:]
                z = str(row[col_zone]).strip() if col_zone < len(row) and row[col_zone] is not None else ""
                if not z or z.lower().startswith("zone wise") or "total" in z.lower() or z.lower() in ["zone", "zone name"] or "ccrs" in z.lower() or "complaints" in z.lower():
                    continue

                r_data = {"row_index": r_idx + 1, "zone": z}
                calc_tot = 0.0

                for i, h in enumerate(headers):
                    if i == col_zone:
                        continue
                    val = _safe_float(row[i]) if i < len(row) else 0.0
                    clean_h = _clean_category_name(h, "water")
                    r_data[clean_h] = _safe_int_or_float(val)
                    if h.lower() not in ["total open", "grand total", "total"]:
                        calc_tot += val

                if "Total Open" not in r_data:
                    r_data["Total Open"] = _safe_int_or_float(calc_tot)

                matrix_rows.append(r_data)

            if matrix_rows:
                return matrix_rows

        # 2. Row-wise format
        if has_problem_col:
            col_zone = next((i for i, h in enumerate(headers) if "zone" in h.lower() or "row labels" in h.lower()), 0)
            col_prob = next((i for i, h in enumerate(headers) if "problem" in h.lower() or "category" in h.lower() or "nature" in h.lower()), 1)
            col_open = next((i for i, h in enumerate(headers) if "open" in h.lower() and "re" not in h.lower() and "grand" not in h.lower()), None)
            if col_open is None:
                col_open = next((i for i, h in enumerate(headers) if "open" in h.lower()), None)

            zone_categories = {}
            all_categories = set()
            current_zone = ""

            for r_idx in range(header_idx + 1, len(all_rows)):
                raw_row = all_rows[r_idx]
                if not raw_row:
                    continue
                row = raw_row[first_non_empty:]
                raw_z = str(row[col_zone]).strip() if col_zone < len(row) and row[col_zone] is not None else ""
                raw_p = str(row[col_prob]).strip() if col_prob < len(row) and row[col_prob] is not None else ""

                if raw_z.lower().startswith("zone wise") or "total" in raw_z.lower() or raw_z.lower() in ["zone", "row labels"] or "ccrs" in raw_z.lower() or "complaints" in raw_z.lower():
                    if "total" in raw_z.lower():
                        current_zone = ""
                    continue

                if raw_z != "" and not raw_z.lower().endswith("total"):
                    current_zone = raw_z

                if not raw_p or raw_p.lower() in ["problem", "row labels"] or "total" in raw_p.lower():
                    continue

                if not current_zone:
                    continue

                cat_name = _clean_category_name(raw_p, "water")
                open_val = _safe_float(row[col_open]) if col_open is not None and col_open < len(row) else 0.0

                if current_zone not in zone_categories:
                    zone_categories[current_zone] = {}
                zone_categories[current_zone][cat_name] = zone_categories[current_zone].get(cat_name, 0.0) + open_val
                all_categories.add(cat_name)

            matrix_rows = []
            for z, cats in zone_categories.items():
                r_data = {"zone": z}
                tot_open = 0.0
                for cat in sorted(list(all_categories)):
                    val = cats.get(cat, 0.0)
                    int_val = _safe_int_or_float(val)
                    r_data[cat] = int_val
                    tot_open += val
                r_data["Total Open"] = _safe_int_or_float(tot_open)
                matrix_rows.append(r_data)

            if matrix_rows:
                return matrix_rows

        # 3. Pivot format (Row Labels, Closed, Open)
        col_label = 0
        for i, h in enumerate(headers):
            if h.lower() in ["row labels", "zone", "problem"]:
                col_label = i
                break
        col_open = next((i for i, h in enumerate(headers) if "open" in h.lower() and "re" not in h.lower() and "grand" not in h.lower()), 1)

        zone_categories = {}
        all_categories = set()
        current_zone = ""

        for r_idx in range(header_idx + 1, len(all_rows)):
            raw_row = all_rows[r_idx]
            if not raw_row:
                continue
            row = raw_row[first_non_empty:]
            lbl = str(row[col_label]).strip() if col_label < len(row) and row[col_label] is not None else ""
            if not lbl or lbl.lower().startswith("zone wise") or "total" in lbl.lower() or lbl.lower() in ["row labels", "zone"] or "ccrs" in lbl.lower() or "complaints" in lbl.lower():
                continue

            if lbl.isupper() and len(lbl) < 20 and not lbl.startswith("WATER"):
                current_zone = lbl
                continue

            if not current_zone:
                continue

            cat_name = _clean_category_name(lbl, "water")
            open_val = _safe_float(row[col_open]) if col_open < len(row) else 0.0

            if current_zone not in zone_categories:
                zone_categories[current_zone] = {}

            zone_categories[current_zone][cat_name] = zone_categories[current_zone].get(cat_name, 0.0) + open_val
            all_categories.add(cat_name)

        matrix_rows = []
        for z, cats in zone_categories.items():
            r_data = {"zone": z}
            tot_open = 0.0
            for cat in sorted(list(all_categories)):
                val = cats.get(cat, 0.0)
                int_val = _safe_int_or_float(val)
                r_data[cat] = int_val
                tot_open += val
            r_data["Total Open"] = _safe_int_or_float(tot_open)
            matrix_rows.append(r_data)

        return matrix_rows
    except Exception as e:
        raise ValueError(f"Error parsing Water Excel file: {str(e)}")
